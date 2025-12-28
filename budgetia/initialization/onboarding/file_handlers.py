# src/initialization/onboarding/file_handlers.py
from abc import ABC, abstractmethod
from dataclasses import dataclass
from typing import Any
from pathlib import Path

from core.google_auth_service import GoogleAuthService


# Interface para retorno padronizado
@dataclass
class AcquisitionResult:
    success: bool
    file_path: str | None = None
    handler_type: str = "unknown"
    requires_ui_action: str | None = (
        None  # Ex: "show_file_uploader", "show_google_oauth"
    )
    error_message: str | None = None


class IFileHandler(ABC):
    """Interface para manipuladores de aquisi√ß√£o de planilha."""

    @abstractmethod
    def can_handle(self, user_input: str) -> bool:
        """Verifica se este handler pode processar a inten√ß√£o do usu√°rio."""
        pass

    @abstractmethod
    def acquire(self, context: dict) -> AcquisitionResult:
        """
        Executa a l√≥gica de aquisi√ß√£o.

        Args:
            context: Dicion√°rio com dados necess√°rios (ex: config_service, uploaded_file, username)

        Returns:
            AcquisitionResult com o resultado da opera√ß√£o.
        """
        pass


class DefaultSpreadsheetHandler(IFileHandler):
    """
    Cria uma planilha nova do zero com estrutura completa.
    Acionado por: "come√ßar do zero", "criar nova", "n√£o tenho".
    """

    def can_handle(self, user_input: str) -> bool:
        # Remove emojis e limpa texto
        clean_input = (
            user_input.lower()
            .replace("üöÄ", "")
            .replace("üì§", "")
            .replace("‚òÅÔ∏è", "")
            .strip()
        )

        keywords = [
            "come√ßar do zero",
            "criar do zero",
            "criar nova",
            "n√£o tenho",
            "criar planilha",
            "nova planilha",
        ]
        result = any(k in clean_input for k in keywords)
        print(
            f"[DEBUG DefaultHandler] Input: '{user_input}' | Clean: '{clean_input}' | Match: {result}"
        )
        return result

    def acquire(self, context: dict) -> AcquisitionResult:
        """Cria uma nova planilha Excel com estrutura completa na pasta do usu√°rio."""
        from pathlib import Path

        from openpyxl import Workbook

        import config

        try:
            # Obt√©m username do contexto
            username = context.get("username")
            if not username:
                return AcquisitionResult(
                    success=False,
                    file_path=None,
                    handler_type="default",
                    error_message="Username n√£o fornecido no contexto",
                )

            # Define caminho na pasta do usu√°rio
            user_dir = Path(config.DATA_DIR) / "users" / username
            user_dir.mkdir(parents=True, exist_ok=True)

            file_path = user_dir / "MinhasFinancas.xlsx"

            print(
                f"[DEBUG DefaultHandler] Criando planilha com estrutura em: {file_path}"
            )

            # Cria workbook com todas as abas e estrutura
            wb = Workbook()
            wb.remove(wb.active)  # Remove aba padr√£o

            # 1. Aba Transa√ß√µes
            ws_transacoes = wb.create_sheet(config.NomesAbas.TRANSACOES)
            transacoes_cols = [
                config.ColunasTransacoes.ID,
                config.ColunasTransacoes.DATA,
                config.ColunasTransacoes.TIPO,
                config.ColunasTransacoes.CATEGORIA,
                config.ColunasTransacoes.DESCRICAO,
                config.ColunasTransacoes.VALOR,
                config.ColunasTransacoes.STATUS,
            ]
            ws_transacoes.append(transacoes_cols)

            # 2. Aba Or√ßamentos
            ws_orcamentos = wb.create_sheet(config.NomesAbas.ORCAMENTOS)
            orcamentos_cols = [
                config.ColunasOrcamentos.ID,
                config.ColunasOrcamentos.CATEGORIA,
                config.ColunasOrcamentos.LIMITE,
                config.ColunasOrcamentos.GASTO,
                config.ColunasOrcamentos.PERCENTUAL,
                config.ColunasOrcamentos.PERIODO,
                config.ColunasOrcamentos.STATUS,
                config.ColunasOrcamentos.OBS,
                config.ColunasOrcamentos.ATUALIZACAO,
            ]
            ws_orcamentos.append(orcamentos_cols)

            # 3. Aba D√≠vidas
            ws_dividas = wb.create_sheet(config.NomesAbas.DIVIDAS)
            dividas_cols = [
                config.ColunasDividas.ID,
                config.ColunasDividas.NOME,
                config.ColunasDividas.VALOR_ORIGINAL,
                config.ColunasDividas.SALDO_DEVEDOR,
                config.ColunasDividas.TAXA_JUROS,
                config.ColunasDividas.PARCELAS_TOTAIS,
                config.ColunasDividas.PARCELAS_PAGAS,
                config.ColunasDividas.VALOR_PARCELA,
                config.ColunasDividas.DATA_PGTO,
                config.ColunasDividas.OBS,
            ]
            ws_dividas.append(dividas_cols)

            # 4. Aba Metas
            ws_metas = wb.create_sheet(config.NomesAbas.METAS)
            metas_cols = [
                config.ColunasMetas.ID,
                config.ColunasMetas.NOME,
                config.ColunasMetas.VALOR_ALVO,
                config.ColunasMetas.VALOR_ATUAL,
                config.ColunasMetas.DATA_ALVO,
                config.ColunasMetas.STATUS,
                config.ColunasMetas.OBS,
            ]
            ws_metas.append(metas_cols)

            # 5. Aba Consultoria
            ws_consultoria = wb.create_sheet(config.NomesAbas.CONSULTORIA_IA)
            consultoria_cols = ["Data", "Pergunta", "Resposta da IA"]
            ws_consultoria.append(consultoria_cols)

            # 6. Aba Perfil
            ws_perfil = wb.create_sheet(config.NomesAbas.PERFIL_FINANCEIRO)
            perfil_cols = ["Campo", "Valor"]
            ws_perfil.append(perfil_cols)
            # Adiciona campos padr√£o vazios
            ws_perfil.append(["Renda Mensal M√©dia", ""])
            ws_perfil.append(["Principal Objetivo", ""])

            # Salva arquivo
            wb.save(str(file_path))
            print(
                f"[DEBUG DefaultHandler] Successfully created structured Excel: {file_path}"
            )

            return AcquisitionResult(
                success=True,
                file_path=str(file_path),
                handler_type="default",
            )

        except Exception as e:
            print(f"[DEBUG DefaultHandler] Error creating file: {e}")
            import traceback

            traceback.print_exc()
            return AcquisitionResult(
                success=False,
                file_path=None,
                handler_type="default",
                error_message=str(e),
            )


class UploadHandler(IFileHandler):
    """
    Gerencia upload de arquivo Excel existente.
    Acionado por: "j√° tenho", "upload", "enviar arquivo", "tenho planilha".
    """

    def can_handle(self, user_input: str) -> bool:
        clean_input = (
            user_input.lower()
            .replace("üöÄ", "")
            .replace("üì§", "")
            .replace("‚òÅÔ∏è", "")
            .strip()
        )
        keywords = [
            "j√° tenho",
            "upload",
            "fazer upload",
            "enviar arquivo",
            "tenho planilha",
            "arquivo excel",
        ]
        result = any(k in clean_input for k in keywords)
        print(
            f"[DEBUG UploadHandler] Input: '{user_input}' | Clean: '{clean_input}' | Match: {result}"
        )
        return result

    def acquire(self, context: dict) -> AcquisitionResult:
        uploaded_file = context.get("uploaded_file")

        # Se n√£o tem arquivo no contexto, solicita UI action
        if not uploaded_file:
            return AcquisitionResult(
                success=False,
                file_path=None,
                handler_type="upload",
                requires_ui_action="show_file_uploader",
                error_message="Aguardando upload do arquivo",
            )

        # Se tem arquivo, salva
        save_dir = context.get("save_dir", ".")
        try:
            save_path = Path(save_dir) / uploaded_file.name
            with open(save_path, "wb") as f:
                f.write(uploaded_file.getbuffer())

            print(f"[DEBUG UploadHandler] File saved: {save_path}")
            return AcquisitionResult(
                success=True,
                file_path=str(save_path),
                handler_type="upload",
            )
        except Exception as e:
            print(f"[DEBUG UploadHandler] Error: {e}")
            return AcquisitionResult(
                success=False,
                file_path=None,
                handler_type="upload",
                error_message=f"Erro ao salvar arquivo: {e}",
            )


class GoogleSheetsHandler(IFileHandler):
    """
    Gerencia conex√£o com Google Sheets.
    Acionado por: "google", "drive", "sheets", "nuvem", "online".
    """

    def __init__(self, auth_service: GoogleAuthService):
        self.auth_service = auth_service

    def can_handle(self, user_input: str) -> bool:
        # Check for direct URL
        if "docs.google.com/spreadsheets" in user_input:
            print(f"[DEBUG GoogleHandler] Input is Google Sheets URL: {user_input}")
            return True

        clean_input = (
            user_input.lower()
            .replace("üöÄ", "")
            .replace("üì§", "")
            .replace("‚òÅÔ∏è", "")
            .strip()
        )
        keywords = [
            "google",
            "drive",
            "sheets",
            "nuvem",
            "online",
            "sele√ß√£o de planilha",
        ]
        result = any(k in clean_input for k in keywords)
        print(
            f"[DEBUG GoogleHandler] Input: '{user_input}' | Clean: '{clean_input}' | Match: {result}"
        )
        return result

    def acquire(self, context: dict) -> AcquisitionResult:
        # 1. FIRST: Se a URL foi passada via texto (Chat) ou contexto expl√≠cito, finaliza!
        selected_url = context.get("google_file_url")
        user_input = context.get("user_input_text", "")
        
        # Se o input do usu√°rio parece uma URL de planilha, usamos ele.
        if "docs.google.com/spreadsheets" in user_input:
            selected_url = user_input.strip()

        if selected_url:
            # Limpa c√≥digo de auth antigo para evitar reuso
            if "google_auth_code" in context:
                del context["google_auth_code"]
                
            return AcquisitionResult(
                success=True,
                file_path=selected_url,
                handler_type="google",
            )
        
        # 2. Trata o callback do Google (se o 'code' estiver no contexto)
        auth_code = context.get("google_auth_code")
        redirect_uri = context.get("redirect_uri")  # ‚Üê NOVO: Pega URI do contexto

        if auth_code:
            try:
                self.auth_service.exchange_code_for_tokens(
                    auth_code,
                    redirect_uri=redirect_uri # ‚Üê Passa para o servi√ßo
                )  # Troca c√≥digo por tokens
                # Tokens salvos. Avan√ßa para sele√ß√£o de arquivo.
                return AcquisitionResult(
                    success=False,
                    handler_type="google",
                    requires_ui_action="show_file_selection",  # <--- Pr√≥ximo passo
                    error_message="Autentica√ß√£o conclu√≠da. Por favor, selecione sua planilha.",
                )
            except Exception as e:
                return AcquisitionResult(
                    success=False,
                    handler_type="google",
                    requires_ui_action="show_google_oauth",
                    error_message=f"Erro ao trocar o c√≥digo por tokens. Tente novamente: {e}",
                )

        # 3. Se j√° estamos autenticados (ou ap√≥s a troca), mas a URL n√£o foi selecionada
        # Ou se o usu√°rio clicou para conectar e j√° tinha credenciais antigas v√°lidas.
        if self.auth_service.get_user_credentials():
            return AcquisitionResult(
                success=False,
                handler_type="google",
                requires_ui_action="show_file_selection",  # <--- Pr√≥ximo passo
                error_message="Por favor, selecione sua planilha do Google Drive.",
            )

        # 4. Se n√£o tem credenciais nem URL, solicita OAuth (in√≠cio do fluxo)
        return AcquisitionResult(
            success=False,
            file_path=None,
            handler_type="google",
            requires_ui_action="show_google_oauth",  # <--- Inicia o fluxo
            error_message="Aguardando sele√ß√£o do Google Sheets",
        )
