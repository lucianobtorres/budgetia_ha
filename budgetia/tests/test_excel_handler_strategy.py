# Em: tests/test_excel_handler_strategy.py

import io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

# Importa as classes que vamos testar
from BudgetIA.src import config
from finance.storage.excel_storage_handler import ExcelStorageHandler
from finance.strategies.custom_json_strategy import CustomJsonStrategy


@pytest.fixture
def fake_excel_user_buffer() -> bytes:
    """Cria um arquivo Excel com formato de usuário (colunas 'estranhas') em memória."""
    df_usuario = pd.DataFrame(
        {
            "Data Operação": ["2024-10-01", "2024-10-03"],
            "Histórico": ["Salário", "Conta de Luz"],
            "Montante": [5000.00, -180.50],  # Usa negativo para despesa
        }
    )
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_usuario.to_excel(writer, sheet_name="Extrato_Banco", index=False)
    buffer.seek(0)
    return buffer.getvalue()


@pytest.fixture
def user_map() -> dict:
    """Simula o JSON de mapeamento que o 💰_BudgetIA.py salvaria."""
    return {
        "aba_transacoes": "Extrato_Banco",
        "colunas": {
            # "Coluna do Usuário": "Nossa Coluna Padrão"
            "Data Operação": "Data",
            "Montante": "Valor",
            "Histórico": "Descricao",
        },
        "transform_valor_negativo": True,  # Importante!
    }


def test_strategy_load_and_unmap(
    tmp_path: Path, fake_excel_user_buffer: bytes, user_map: dict
) -> None:
    """
    Testa o ciclo completo:
    1. Ler (Load) um Excel de usuário -> map_transactions
    2. Salvar (Save) no nosso formato -> unmap_transactions
    """
    file_path = str(tmp_path / "planilha_teste_ciclo.xlsx")

    # --- ARRANGE (Salvando o arquivo "fake" do usuário no disco) ---
    with open(file_path, "wb") as f:
        f.write(fake_excel_user_buffer)

    # --- ETAPA 1: TESTE DE LEITURA (map_transactions) ---

    # 1. Cria a Estratégia de Leitura
    strategy = CustomJsonStrategy(config.LAYOUT_PLANILHA, user_map)
    # 2. Cria o Handler
    handler_leitura = ExcelStorageHandler(file_path=file_path)
    # 3. Executa o Load (que usa strategy.map_transactions)
    dfs_interno, _ = handler_leitura.load_sheets(config.LAYOUT_PLANILHA, strategy)

    # --- ASSERT LEITURA ---
    df_transacoes = dfs_interno[config.NomesAbas.TRANSACOES]

    # Verifica se as colunas foram renomeadas
    assert config.ColunasTransacoes.VALOR in df_transacoes.columns
    assert config.ColunasTransacoes.DESCRICAO in df_transacoes.columns
    # Verifica se a transformação de negativo funcionou
    assert df_transacoes.iloc[0][config.ColunasTransacoes.VALOR] == 5000.00
    assert (
        df_transacoes.iloc[0][config.ColunasTransacoes.TIPO]
        == config.ValoresTipo.RECEITA
    )
    assert df_transacoes.iloc[1][config.ColunasTransacoes.VALOR] == 180.50
    assert (
        df_transacoes.iloc[1][config.ColunasTransacoes.TIPO]
        == config.ValoresTipo.DESPESA
    )
    # Verifica se colunas internas foram adicionadas
    assert config.ColunasTransacoes.ID in df_transacoes.columns
    assert config.ColunasTransacoes.STATUS in df_transacoes.columns

    # --- ETAPA 2: TESTE DE ESCRITA (unmap_transactions) ---

    # 1. Modifica o DataFrame interno (simulando a IA)
    df_transacoes.loc[
        df_transacoes[config.ColunasTransacoes.DESCRICAO] == "Conta de Luz",
        config.ColunasTransacoes.STATUS,
    ] = "Verificado"

    # 2. Cria o Handler de Escrita
    handler_escrita = ExcelStorageHandler(file_path=file_path)
    # 3. Executa o Save (que usa strategy.unmap_transactions)
    handler_escrita.save_sheets(dfs_interno, strategy)

    # --- ASSERT ESCRITA ---

    # 1. Re-lê o arquivo salvo (agora com pandas, sem estratégia)
    wb = load_workbook(file_path)
    # 2. Verifica se a aba foi salva com o NOME DO USUÁRIO
    assert "Extrato_Banco" in wb.sheetnames
    assert (
        config.NomesAbas.TRANSACOES not in wb.sheetnames
    )  # Prova que não salvou com nosso nome

    # 3. Lê os dados da aba salva
    df_salvo = pd.read_excel(file_path, sheet_name="Extrato_Banco")

    # 4. Verifica se as colunas internas ("ID", "Status") foram REMOVIDAS
    assert config.ColunasTransacoes.STATUS not in df_salvo.columns
    assert config.ColunasTransacoes.ID not in df_salvo.columns

    # 5. Verifica se as colunas voltaram ao NOME DO USUÁRIO
    assert "Montante" in df_salvo.columns
    assert "Histórico" in df_salvo.columns
    assert "Valor" not in df_salvo.columns  # Prova que foi renomeado

    # 6. Verifica se a transformação INVERSA de negativo funcionou
    assert df_salvo.iloc[0]["Montante"] == 5000.00
    assert df_salvo.iloc[1]["Montante"] == -180.50
